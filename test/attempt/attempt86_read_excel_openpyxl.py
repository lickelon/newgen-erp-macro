"""엑셀 파일 구조 확인 스크립트"""
import pandas as pd
import sys

sys.stdout.reconfigure(encoding='utf-8')

try:
    # 엑셀 파일 읽기
    print("=" * 70)
    print("테스트 임직원.xlsx 읽기")
    print("=" * 70)

    # openpyxl로 직접 읽기 시도 (data_only=True)
    from openpyxl import load_workbook

    print("\n시도 1: openpyxl (data_only=True, read_only=True)")
    wb = load_workbook('테스트 임직원.xlsx', data_only=True, read_only=True)
    ws = wb.active

    # 데이터를 리스트로 변환
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # pandas DataFrame으로 변환
    df = pd.DataFrame(data[1:], columns=data[0])  # 첫 행을 컬럼으로 사용

    print(f"\n✓ 파일 로드 성공!")
    print(f"  행 수: {len(df)}")
    print(f"  열 수: {len(df.columns)}")

    print(f"\n컬럼 목록:")
    for i, col in enumerate(df.columns):
        print(f"  {i+1}. {col}")

    print(f"\n첫 5행 데이터:")
    print(df.head().to_string())

    print(f"\n데이터 타입:")
    print(df.dtypes)

    print(f"\n기본 통계:")
    print(df.describe())

except Exception as e:
    print(f"\n❌ 오류 발생: {e}")
    import traceback
    traceback.print_exc()
