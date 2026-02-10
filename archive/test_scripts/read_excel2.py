import sys
from openpyxl import load_workbook

# UTF-8 출력 설정
sys.stdout.reconfigure(encoding='utf-8')

wb = load_workbook('연말정산.xls')

# 모든 시트 이름
print("시트 목록:", wb.sheetnames)
print()

# 첫 번째 시트
ws = wb.active
print(f"첫 번째 시트: {ws.title}")
print(f"총 {ws.max_row}행, {ws.max_column}열\n")

# 헤더
headers = [cell for cell in next(ws.iter_rows(values_only=True))]
print("컬럼:", headers)
print()

# 처음 5개 행
print("처음 5개 데이터:")
for idx, row in enumerate(list(ws.iter_rows(values_only=True))[1:6], 1):
    print(f"\n행 {idx}:")
    for header, value in zip(headers, row):
        if value is not None:
            print(f"  {header}: {value}")
