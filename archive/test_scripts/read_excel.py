import sys
from openpyxl import load_workbook
import json

# UTF-8 출력 설정
sys.stdout.reconfigure(encoding='utf-8')

wb = load_workbook('테스트 데이터.xlsx')
ws = wb.active

print(f"총 {ws.max_row}행, {ws.max_column}열\n")

# 헤더
headers = [cell for cell in next(ws.iter_rows(values_only=True))]
print("컬럼:", headers)
print()

# 처음 5개 행
print("처음 5개 데이터:")
for idx, row in enumerate(list(ws.iter_rows(values_only=True))[1:6], 1):
    print(f"\n행 {idx}:")
    for header, value in zip(headers, row):
        print(f"  {header}: {value}")
